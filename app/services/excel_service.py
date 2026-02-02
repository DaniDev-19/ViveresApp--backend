import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO


class ExcelService:
    def generate_report(self, data: list, headers: list, title: str = "Reporte"):
        """
        Genera un archivo Excel estilizado a partir de una lista de diccionarios.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title

        # Estilo del Encabezado
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F46E5", end_color="4F46E5", fill_type="solid"
        )  # Indio
        alignment = Alignment(horizontal="center", vertical="center")

        # Escribir Encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment

        # Escribir Datos
        for row_num, row_data in enumerate(data, 2):
            for col_num, key in enumerate(row_data.keys(), 1):
                ws.cell(row=row_num, column=col_num, value=str(row_data[key]))

        # Ajustar ancho de columnas automáticamente
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer


excel_service = ExcelService()
