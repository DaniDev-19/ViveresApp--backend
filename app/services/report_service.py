from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict, Any
from io import BytesIO
from datetime import datetime
from app.core.config import settings

class ReportService:
    def _get_header_style(self):
        return ParagraphStyle(
            'Header',
            parent=getSampleStyleSheet()['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4F46E5'), # Indigo 600
            alignment=TA_CENTER,
            spaceAfter=20
        )

    def _get_sub_header_style(self):
         return ParagraphStyle(
            'SubHeader',
            parent=getSampleStyleSheet()['Normal'],
            fontSize=12,
            textColor=colors.grey,
            alignment=TA_CENTER,
            spaceAfter=30
        )

    def generate_sales_pdf(self, sales_data: List[Dict[str, Any]], date_range: str) -> BytesIO:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
        elements = []
        
        elements.append(Paragraph(settings.BUSINESS_NAME.upper(), self._get_header_style()))
        elements.append(Paragraph(f"Reporte de Ventas Detallado", self._get_sub_header_style()))
        elements.append(Paragraph(f"Periodo: {date_range} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", self._get_sub_header_style()))

        # Columns: ID, Fecha, Cliente, Deliv, IVA, Ganancia, Total BS, Total USD
        headers = ["ID", "Fecha", "Cliente", "Deliv.", "IVA", "Ganancia", "Total (BS)", "Total ($)"]
        data = [headers]
        
        # Totals
        sum_delivery = 0
        sum_tax = 0
        sum_profit = 0
        sum_bs = 0
        sum_usd = 0

        for sale in sales_data:
            d = sale.get("delivery", 0)
            t = sale.get("tax", 0)
            p = sale.get("profit", 0)
            bs = sale.get("total_bs", 0)
            usd = sale.get("total", 0)

            data.append([
                str(sale.get("id", "")),
                sale.get("date", ""),
                sale.get("customer", "Cliente Casual")[:15],
                f"${d:.2f}",
                f"${t:.2f}",
                f"${p:.2f}",
                f"{bs:,.2f}",
                f"${usd:.2f}",
            ])
            
            sum_delivery += d
            sum_tax += t
            sum_profit += p
            sum_bs += bs
            sum_usd += usd

        # Summary Row
        data.append([
            "TOTALES", "", "",
            f"${sum_delivery:.2f}",
            f"${sum_tax:.2f}",
            f"${sum_profit:.2f}",
            f"{sum_bs:,.2f}",
            f"${sum_usd:.2f}"
        ])

        col_widths = [30, 80, 110, 50, 50, 60, 90, 80]

        t = Table(data, colWidths=col_widths)
        
        # Styles
        styles = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#F3F4F6')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')), # Extended to -1 (footer)
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F9FAFB')]),
            
            # Footer Style (Last Row)
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E0E7FF')), # Light Indigo
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.black),
            ('ALIGN', (0, -1), (0, -1), 'RIGHT'), # Align "TOTALES" to right of the spanned cell
            ('TOPPADDING', (0, -1), (-1, -1), 12),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 12),
            ('SPAN', (0, -1), (2, -1)), # Merge first 3 cols for "TOTALES" label (UPPERCASE)
        ]
        
        t.setStyle(TableStyle(styles))
        elements.append(t)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_sales_excel(self, sales_data: List[Dict[str, Any]]) -> BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Ventas"

        ws.merge_cells('A1:K1')
        ws['A1'] = settings.BUSINESS_NAME.upper()
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        headers = ["ID", "Fecha", "Cliente", "Cajero", "Items", "Delivery", "IVA", "Ganancia", "Total (BS)", "Total (USD)", "Estado"]
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15

        ws.column_dimensions['E'].width = 40 
        ws.column_dimensions['C'].width = 25

        last_row = 3
        for row_num, sale in enumerate(sales_data, 3):
            try:
                ws.cell(row=row_num, column=1, value=sale.get("id"))
                ws.cell(row=row_num, column=2, value=str(sale.get("date")))
                ws.cell(row=row_num, column=3, value=str(sale.get("customer", "")))
                ws.cell(row=row_num, column=4, value=str(sale.get("cashier", "")))
                ws.cell(row=row_num, column=5, value=str(sale.get("items", "")))
                
                # Numeric fields
                ws.cell(row=row_num, column=6, value=sale.get("delivery", 0)).number_format = '"$"#,##0.00'
                ws.cell(row=row_num, column=7, value=sale.get("tax", 0)).number_format = '"$"#,##0.00'
                ws.cell(row=row_num, column=8, value=sale.get("profit", 0)).number_format = '"$"#,##0.00'
                ws.cell(row=row_num, column=9, value=sale.get("total_bs", 0)).number_format = '#,##0.00'
                ws.cell(row=row_num, column=10, value=sale.get("total", 0)).number_format = '"$"#,##0.00'
                
                ws.cell(row=row_num, column=11, value=str(sale.get("status", "")))

                for i in range(1, 12):
                    ws.cell(row=row_num, column=i).border = border
                
                last_row = row_num
            except Exception as e:
                print(f"Error processing row {row_num}: {e}")
                continue
        
        # Totals Row
        total_row = last_row + 1
        ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
        # Sum Delivery (F) to Total USD (J)
        cols_to_sum = ['F', 'G', 'H', 'I', 'J']
        
        for col in cols_to_sum:
            cell = ws.cell(row=total_row, column=openpyxl.utils.column_index_from_string(col))
            cell.value = f"=SUM({col}3:{col}{last_row})"
            cell.font = Font(bold=True)
            if col == 'I':
                cell.number_format = '#,##0.00'
            else:
                cell.number_format = '"$"#,##0.00'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_inventory_pdf(self, products_data: List[Dict[str, Any]], report_type: str = "standard") -> BytesIO:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        
        elements.append(Paragraph(settings.BUSINESS_NAME.upper(), self._get_header_style()))
        
        if report_type == "code_name":
            title_text = "Reporte de Catálogo (Código y Nombre)"
            headers = ["Código", "Producto"]
            col_widths = [150, 402]
        elif report_type == "prices":
            title_text = "Reporte de Precios de Productos"
            headers = ["Código", "Producto", "Precio Oferta", "Precio Final"]
            col_widths = [120, 252, 90, 90]
        else:
            title_text = "Reporte de Inventario General"
            headers = ["ID", "Código", "Producto", "Stock", "Precio", "Valor Total"]
            col_widths = [30, 90, 190, 60, 80, 80]

        elements.append(Paragraph(title_text, self._get_sub_header_style()))
        elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", self._get_sub_header_style()))

        data = [headers]
        total_value = 0

        for prod in products_data:
            if report_type == "code_name":
                data.append([
                    prod.get("barcode", "N/A"),
                    prod.get("name", "Producto")[:50]
                ])
            elif report_type == "prices":
                off_price = prod.get("offer_price", 0.0)
                off_price_str = f"${off_price:.2f}" if off_price > 0 else "-"
                data.append([
                    prod.get("barcode", "N/A"),
                    prod.get("name", "Producto")[:35],
                    off_price_str,
                    f"${prod.get('price', 0):.2f}"
                ])
            else:
                val = prod.get("stock", 0) * prod.get("price", 0)
                data.append([
                    str(prod.get("id", "")),
                    prod.get("barcode", "N/A")[:15],
                    prod.get("name", "Producto")[:30],
                    str(prod.get("stock", 0)),
                    f"${prod.get('price', 0):.2f}",
                    f"${val:.2f}"
                ])
                total_value += val

        t = Table(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')), 
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('TOPPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ECFDF5')]),
        ]))
        elements.append(t)
        
        if report_type == "standard":
            elements.append(Spacer(1, 24))
            elements.append(Paragraph(f"Valor Total Inventario: ${total_value:.2f}", ParagraphStyle('Total', parent=getSampleStyleSheet()['Heading2'], alignment=TA_RIGHT, textColor=colors.HexColor('#065F46'))))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_inventory_excel(self, products_data: List[Dict[str, Any]], report_type: str = "standard") -> BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"

        if report_type == "code_name":
            headers = ["Código", "Producto"]
            max_col = 2
        elif report_type == "prices":
            headers = ["Código", "Producto", "Precio Oferta", "Precio Final"]
            max_col = 4
        else:
            headers = ["ID", "Código", "Producto", "Stock", "Costo", "Precio", "Margen", "Valor Total"]
            max_col = 8

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        ws.cell(row=1, column=1, value=settings.BUSINESS_NAME.upper())
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15

        if report_type in ["code_name", "prices"]:
            ws.column_dimensions['B'].width = 40  # Name
        else:
            ws.column_dimensions['C'].width = 40  # Name

        for row_num, prod in enumerate(products_data, 3):
            if report_type == "code_name":
                ws.cell(row=row_num, column=1, value=prod.get("barcode", "N/A"))
                ws.cell(row=row_num, column=2, value=prod.get("name"))
            elif report_type == "prices":
                off_price = prod.get("offer_price", 0.0)
                ws.cell(row=row_num, column=1, value=prod.get("barcode", "N/A"))
                ws.cell(row=row_num, column=2, value=prod.get("name"))
                ws.cell(row=row_num, column=3, value=off_price if off_price > 0 else "-").number_format = '"$"#,##0.00' if off_price > 0 else '@'
                ws.cell(row=row_num, column=4, value=prod.get("price")).number_format = '"$"#,##0.00'
            else:
                val = prod.get("stock", 0) * prod.get("price", 0)
                ws.cell(row=row_num, column=1, value=prod.get("id"))
                ws.cell(row=row_num, column=2, value=prod.get("barcode", "N/A"))
                ws.cell(row=row_num, column=3, value=prod.get("name"))
                ws.cell(row=row_num, column=4, value=prod.get("stock"))
                ws.cell(row=row_num, column=5, value=prod.get("cost")).number_format = '"$"#,##0.00'
                ws.cell(row=row_num, column=6, value=prod.get("price")).number_format = '"$"#,##0.00'
                ws.cell(row=row_num, column=7, value=prod.get("margin"))
                ws.cell(row=row_num, column=8, value=val).number_format = '"$"#,##0.00'

            for i in range(1, max_col + 1):
                ws.cell(row=row_num, column=i).border = border

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_labels_pdf(self, products_to_print: List[Dict[str, Any]]) -> BytesIO:
        buffer = BytesIO()
        c = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        cols = 3
        rows = 8
        margin_x = 20
        margin_y = 20
        
        col_width = (width - 2 * margin_x) / cols
        row_height = (height - 2 * margin_y) / rows
        
        x_idx = 0
        y_idx = 0

        name_style = ParagraphStyle(
            'LabelProductName',
            parent=getSampleStyleSheet()['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            leading=11,
            alignment=TA_CENTER,
            textColor=colors.black
        )

        for item in products_to_print:
            x = margin_x + x_idx * col_width
            y = height - margin_y - (y_idx + 1) * row_height
            
            # Border
            c.setStrokeColor(colors.lightgrey)
            c.setDash(1, 2)
            c.rect(x + 5, y + 5, col_width - 10, row_height - 10)
            c.setDash([])
            c.setStrokeColor(colors.black)

            # Content
            # 1. Product Name (Top) - Wrapped using Paragraph to support multi-line wrap cleanly
            name = item.get("name", "")
            if len(name) > 60:
                name = name[:57] + "..."

            p = Paragraph(name, name_style)
            p_width, p_height = p.wrap(col_width - 20, row_height - 30)
            
            # Draw name at the top inside cell margins
            p.drawOn(c, x + 10, y + row_height - p_height - 10)
            
            # 2. Price (Dynamic position based on remaining space to prevent overlap)
            # Center of the remaining space between the bottom of name and top of business footer (y + 20)
            remaining_top = row_height - p_height - 10
            price_y_offset = (remaining_top + 20) / 2
            
            c.setFont("Helvetica-Bold", 26)
            c.drawCentredString(x + col_width / 2, y + price_y_offset - 10, f"${item.get('price', 0):.2f}")
            
            # 3. Footer (Bottom)
            c.setFont("Helvetica-Oblique", 9)
            c.setFillColor(colors.HexColor('#4F46E5'))
            c.drawCentredString(x + col_width / 2, y + 15, settings.BUSINESS_NAME.upper())

            x_idx += 1
            if x_idx >= cols:
                x_idx = 0
                y_idx += 1
                if y_idx >= rows:
                    c.showPage()
                    y_idx = 0
                    
        c.save()
        buffer.seek(0)
        return buffer
        
    def generate_monthly_growth_pdf(self, monthly_data: List[Dict[str, Any]], year: int) -> BytesIO:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        
        elements.append(Paragraph(settings.BUSINESS_NAME.upper(), self._get_header_style()))
        elements.append(Paragraph(f"Reporte de Crecimiento y Ganancia Mensual - Año {year}", self._get_sub_header_style()))
        elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", self._get_sub_header_style()))
        elements.append(Spacer(1, 15))
        
        headers = ["Mes", "Ventas USD", "Ganancia USD", "Ventas BCV", "Ganancia BCV", "Ventas USDT", "Ganancia USDT"]
        rows = []
        
        total_sales_usd = 0.0
        total_profit_usd = 0.0
        
        for m in monthly_data:
            rows.append([
                m.get("month_name", ""),
                f"${m.get('total_usd', 0.0):,.2f}",
                f"${m.get('profit_usd', 0.0):,.2f}",
                f"Bs. {m.get('total_bs_bcv', 0.0):,.2f}",
                f"Bs. {m.get('profit_bs_bcv', 0.0):,.2f}",
                f"Bs. {m.get('total_bs_usdt', 0.0):,.2f}",
                f"Bs. {m.get('profit_bs_usdt', 0.0):,.2f}",
            ])
            total_sales_usd += m.get('total_usd', 0.0)
            total_profit_usd += m.get('profit_usd', 0.0)
            
        total_sales_bs_bcv = sum(m.get('total_bs_bcv', 0.0) for m in monthly_data)
        total_profit_bs_bcv = sum(m.get('profit_bs_bcv', 0.0) for m in monthly_data)
        total_sales_bs_usdt = sum(m.get('total_bs_usdt', 0.0) for m in monthly_data)
        total_profit_bs_usdt = sum(m.get('profit_bs_usdt', 0.0) for m in monthly_data)
        
        rows.append([
            "TOTAL ANUAL",
            f"${total_sales_usd:,.2f}",
            f"${total_profit_usd:,.2f}",
            f"Bs. {total_sales_bs_bcv:,.2f}",
            f"Bs. {total_profit_bs_bcv:,.2f}",
            f"Bs. {total_sales_bs_usdt:,.2f}",
            f"Bs. {total_profit_bs_usdt:,.2f}",
        ])
        
        col_widths = [72, 75, 75, 85, 85, 80, 80]
        
        t = Table([headers] + rows, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F9FAFB')]),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#EEF2F6')),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 8),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
            ('TOPPADDING', (0, 1), (-1, -1), 5),
        ]))
        
        elements.append(t)
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_rankings_pdf(self, rankings_data: Dict[str, Any], date_range: str) -> BytesIO:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []
        
        elements.append(Paragraph(settings.BUSINESS_NAME.upper(), self._get_header_style()))
        elements.append(Paragraph("Reporte de Rankings y Rendimiento", self._get_sub_header_style()))
        elements.append(Paragraph(f"Periodo: {date_range} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", self._get_sub_header_style()))
        
        section_style = ParagraphStyle(
            'SectionHeader',
            parent=getSampleStyleSheet()['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4F46E5'),
            spaceBefore=15,
            spaceAfter=8,
            keepWithNext=True
        )
        
        def make_table(headers, rows, col_widths, bg_color):
            table_data = [headers] + rows
            t = Table(table_data, colWidths=col_widths)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(bg_color)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
                ('TOPPADDING', (0, 1), (-1, -1), 5),
            ]))
            return t

        # 1. Líderes de Ventas
        elements.append(Paragraph("Líderes de Ventas (Usuarios / Cajeros)", section_style))
        users_headers = ["ID", "Usuario", "Transacciones", "Total Generado ($)"]
        users_rows = [
            [str(u.get("id", "")), u.get("username", ""), str(u.get("transactions", 0)), f"${u.get('total_amount', 0.0):,.2f}"]
            for u in rankings_data.get("top_users", [])
        ]
        if not users_rows:
            users_rows = [["-", "No hay datos", "-", "-"]]
        elements.append(make_table(users_headers, users_rows, [50, 150, 150, 202], '#4F46E5'))
        elements.append(Spacer(1, 15))

        # 2. Productos Más Vendidos
        elements.append(Paragraph("Top 10 Productos Más Vendidos (Unidades)", section_style))
        prod_headers = ["ID Producto", "Nombre del Producto", "Unidades Vendidas"]
        prod_rows = [
            [str(p.get("id", "")), p.get("name", ""), f"{p.get('value', 0):,.0f} uds"]
            for p in rankings_data.get("top_products", [])
        ]
        if not prod_rows:
            prod_rows = [["-", "No hay datos", "-"]]
        elements.append(make_table(prod_headers, prod_rows, [80, 322, 150], '#10B981'))
        elements.append(Spacer(1, 15))

        # 3. Productos de Baja Rotación
        elements.append(Paragraph("Productos de Baja Rotación (Menos Vendidos)", section_style))
        low_prod_headers = ["ID Producto", "Nombre del Producto", "Unidades Vendidas"]
        low_prod_rows = [
            [str(p.get("id", "")), p.get("name", ""), f"{p.get('value', 0):,.0f} uds"]
            for p in rankings_data.get("low_products", [])
        ]
        if not low_prod_rows:
            low_prod_rows = [["-", "No hay datos", "-"]]
        elements.append(make_table(low_prod_headers, low_prod_rows, [80, 322, 150], '#EF4444'))
        elements.append(Spacer(1, 15))

        # 4. Mejores Clientes
        elements.append(Paragraph("Ranking de Mejores Clientes (Mayor Compra)", section_style))
        cust_headers = ["ID Cliente", "Nombre", "Total Pedidos", "Total Gastado ($)"]
        cust_rows = [
            [str(c.get("id", "")), c.get("name", ""), str(c.get("orders", 0)), f"${c.get('amount', 0.0):,.2f}"]
            for c in rankings_data.get("top_customers", [])
        ]
        if not cust_rows:
            cust_rows = [["-", "No hay datos", "-", "-"]]
        elements.append(make_table(cust_headers, cust_rows, [80, 222, 100, 150], '#F59E0B'))
        elements.append(Spacer(1, 15))

        # 5. Desempeño de Proveedores
        elements.append(Paragraph("Desempeño de Proveedores", section_style))
        prov_headers = ["Proveedor", "Órdenes Totales", "Órdenes Completadas", "Fiabilidad"]
        prov_rows = [
            [p.get("name", ""), str(p.get("total_orders", 0)), str(p.get("completed_orders", 0)), f"{p.get('reliability', 0.0):.1f}%"]
            for p in rankings_data.get("top_providers", [])
        ]
        if not prov_rows:
            prov_rows = [["-", "No hay datos", "-", "-"]]
        elements.append(make_table(prov_headers, prov_rows, [222, 100, 130, 100], '#8B5CF6'))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_cash_close_pdf(self, cash_data: Dict[str, Any], date_str: str) -> BytesIO:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []

        # Header
        elements.append(Paragraph(settings.BUSINESS_NAME.upper(), self._get_header_style()))
        elements.append(Paragraph("Reporte de Arqueo y Cierre de Caja", self._get_sub_header_style()))
        elements.append(Paragraph(f"Fecha de Cierre: {date_str} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", self._get_sub_header_style()))
        elements.append(Spacer(1, 15))

        section_style = ParagraphStyle(
            'SectionHeaderClose',
            parent=getSampleStyleSheet()['Heading2'],
            fontSize=13,
            textColor=colors.HexColor('#1E293B'),
            spaceBefore=12,
            spaceAfter=6,
            keepWithNext=True
        )

        # 1. Resumen General
        elements.append(Paragraph("1. Resumen General de Operaciones", section_style))
        
        metrics_headers = ["Transacciones", "Facturación Total ($)", "Total IVA ($)", "Ganancia Neta ($)"]
        metrics_row = [
            str(cash_data.get("total_sales_count", 0)),
            f"${cash_data.get('total_revenue_usd', 0.0):,.2f}",
            f"${cash_data.get('total_tax_usd', 0.0):,.2f}",
            f"${cash_data.get('total_profit_usd', 0.0):,.2f}",
        ]
        
        t_metrics = Table([metrics_headers, metrics_row], colWidths=[130, 130, 130, 130])
        t_metrics.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#334155')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#F8FAFC')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 12),
            ('TEXTCOLOR', (3, 1), (3, 1), colors.HexColor('#10B981')),
        ]))
        
        elements.append(t_metrics)
        elements.append(Spacer(1, 20))

        # 2. Desglose por Método de Pago
        elements.append(Paragraph("2. Desglose de Ventas por Método de Pago", section_style))
        
        pm_headers = ["Método de Pago", "Cant. Operaciones", "Monto Recaudado", "Monto en USD Eq."]
        pm_rows = []
        for pm in cash_data.get("payment_breakdown", []):
            cur = pm.get("currency", "USD")
            sym = "Bs. " if cur == "VES" else "$"
            pm_rows.append([
                pm.get("name", ""),
                str(pm.get("count", 0)),
                f"{sym}{pm.get('amount', 0.0):,.2f}",
                f"${pm.get('amount_usd', 0.0):,.2f}"
            ])
            
        if not pm_rows:
            pm_rows.append(["Sin transacciones en la fecha", "-", "-", "-"])

        t_pm = Table([pm_headers] + pm_rows, colWidths=[180, 100, 120, 120])
        t_pm.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        elements.append(t_pm)
        elements.append(Spacer(1, 25))

        # 3. Conciliación y Firmas
        elements.append(Paragraph("3. Control de Auditoría y Conciliación", section_style))
        
        audit_text = (
            "Este reporte resume el arqueo de los fondos digitales y en efectivo recibidos durante la jornada laboral. "
            "El cajero y el supervisor de la tienda deben firmar a continuación para dar conformidad al cuadre físico realizado contra los totales sugeridos por este sistema."
        )
        elements.append(Paragraph(audit_text, ParagraphStyle('AuditStyle', parent=getSampleStyleSheet()['Normal'], fontSize=9, textColor=colors.HexColor('#6B7280'), spaceAfter=50)))
        
        # Signatures
        sig_data = [
            ["_____________________________________", "_____________________________________"],
            ["Firma Cajero", "Firma Supervisor / Administrador"],
            ["Nombre:", "Nombre:"]
        ]
        t_sig = Table(sig_data, colWidths=[260, 260])
        t_sig.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#374151')),
        ]))
        elements.append(t_sig)

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_low_stock_pdf(self, products_data: List[Dict[str, Any]]) -> BytesIO:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
        elements = []

        # Header
        elements.append(Paragraph(settings.BUSINESS_NAME.upper(), self._get_header_style()))
        elements.append(Paragraph("Reporte de Alerta de Stock y Sugerencia de Reposición", self._get_sub_header_style()))
        elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", self._get_sub_header_style()))
        elements.append(Spacer(1, 15))

        # Summary Metrics
        total_items_low = len(products_data)
        total_investment = sum(p["total_cost_usd"] for p in products_data)
        
        summary_style = ParagraphStyle(
            'SummaryStyle',
            parent=getSampleStyleSheet()['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#1E293B')
        )
        
        summary_text = (
            f"<b>Resumen de Alertas:</b> Se encontraron <b>{total_items_low}</b> productos con stock crítico (bajo o igual al nivel mínimo de seguridad). "
            f"La inversión total estimada para el reabastecimiento sugerido es de <b>${total_investment:,.2f} USD</b>."
        )
        elements.append(Paragraph(summary_text, summary_style))
        elements.append(Spacer(1, 15))

        # Cell styles for word-wrap
        cell_style = ParagraphStyle('CellWrap', parent=getSampleStyleSheet()['Normal'], fontSize=8, leading=10, wordWrap='CJK')
        cell_style_center = ParagraphStyle('CellWrapCenter', parent=getSampleStyleSheet()['Normal'], fontSize=8, leading=10, alignment=TA_CENTER, wordWrap='CJK')

        # Table Headers
        headers = ["Código de Barra", "Nombre del Producto", "Stock Act.", "Stock Mín.", "Sugerido", "Costo U. ($)", "Total Est. ($)", "Proveedor Sugerido"]
        
        rows = []
        for p in products_data:
            rows.append([
                Paragraph(str(p.get("barcode", "")), cell_style_center),
                Paragraph(str(p.get("name", "")), cell_style),
                f"{p.get('stock', 0)}",
                f"{p.get('min_stock', 0)}",
                f"+{p.get('suggested_refill', 0)}",
                f"${p.get('cost_usd', 0.0):,.2f}",
                f"${p.get('total_cost_usd', 0.0):,.2f}",
                Paragraph(str(p.get("provider_name", "Sin Proveedor")), cell_style),
            ])

        col_widths = [95, 210, 50, 50, 55, 75, 75, 130]
        
        t = Table([headers] + rows, colWidths=col_widths)
        
        t_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EF4444')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
        ]
        
        for i, p in enumerate(products_data):
            if p.get("stock", 0) == 0:
                t_style.append(('BACKGROUND', (0, i + 1), (-1, i + 1), colors.HexColor('#FEE2E2')))
            else:
                bg = colors.white if i % 2 == 0 else colors.HexColor('#F9FAFB')
                t_style.append(('BACKGROUND', (0, i + 1), (-1, i + 1), bg))
                
        t.setStyle(TableStyle(t_style))
        elements.append(t)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_deliveries_pdf(self, deliveries_data: List[Dict[str, Any]], summary_stats: Dict[str, Any]) -> BytesIO:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        elements = []

        elements.append(Paragraph(settings.BUSINESS_NAME.upper(), self._get_header_style()))
        elements.append(Paragraph("Reporte de Control de Envíos y Diligencias (Delivery)", self._get_sub_header_style()))
        elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", self._get_sub_header_style()))
        elements.append(Spacer(1, 15))

        section_style = ParagraphStyle(
            'SectionHeaderDeliv',
            parent=getSampleStyleSheet()['Heading2'],
            fontSize=13,
            textColor=colors.HexColor('#1E293B'),
            spaceBefore=12,
            spaceAfter=6,
            keepWithNext=True
        )

        elements.append(Paragraph("1. Resumen de Gastos Operativos de Delivery", section_style))
        
        metrics_headers = ["Viajes Totales", "Viajes Completados", "Viajes Pendientes / En Ruta", "Inversión Operativa Total ($)"]
        metrics_row = [
            str(summary_stats.get("total_trips", 0)),
            str(summary_stats.get("completed_trips", 0)),
            str(summary_stats.get("pending_trips", 0)),
            f"${summary_stats.get('total_cost_usd', 0.0):,.2f}",
        ]
        
        t_metrics = Table([metrics_headers, metrics_row], colWidths=[180, 180, 180, 180])
        t_metrics.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#F8FAFC')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 12),
            ('TEXTCOLOR', (3, 1), (3, 1), colors.HexColor('#4F46E5')),
        ]))
        
        elements.append(t_metrics)
        elements.append(Spacer(1, 20))

        elements.append(Paragraph("2. Detalle de Viajes y Carreras", section_style))

        detail_headers = ["Descripción", "Destino", "Productos", "Proveedor", "Repartidor", "Costo ($)", "Estado", "Registro", "Entrega"]

        cell_style = ParagraphStyle(
            'CellStyle',
            parent=getSampleStyleSheet()['Normal'],
            fontName='Helvetica',
            fontSize=7,
            leading=9,
            wordWrap='CJK',
        )

        rows = []
        for d in deliveries_data:
            rows.append([
                Paragraph(str(d.get("description", "")), cell_style),
                Paragraph(str(d.get("address", "") or "-"), cell_style),
                Paragraph(str(d.get("items_detail", "") or "-"), cell_style),
                Paragraph(str(d.get("provider_name", "Sin Proveedor")), cell_style),
                Paragraph(str(d.get("delivery_user_name", "") or "Sin Asignar"), cell_style),
                f"${d.get('cost_usd', 0.0):,.2f}",
                d.get("status", ""),
                d.get("created_at", "") or "-",
                d.get("completed_at", "") or "-"
            ])
            
        if not rows:
            rows.append(["-", "No se encontraron envíos o viajes en el período", "-", "-", "-", "-", "-", "-", "-"])

        col_widths = [90, 80, 120, 75, 75, 50, 58, 80, 80]
        
        t = Table([detail_headers] + rows, colWidths=col_widths, repeatRows=1)
        t_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (5, 1), (8, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
            ('TOPPADDING', (0, 0), (-1, 0), 5),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
        ]
        
        for i, d in enumerate(deliveries_data):
            st = d.get("status", "")
            if st == "Completado":
                bg = colors.HexColor('#ECFDF5')
            elif st == "Cancelado":
                bg = colors.HexColor('#FEF2F2')
            else:
                bg = colors.white if i % 2 == 0 else colors.HexColor('#F8FAFC')
            t_style.append(('BACKGROUND', (0, i + 1), (-1, i + 1), bg))
            
        t.setStyle(TableStyle(t_style))
        elements.append(t)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_returns_exchanges_pdf(self, returns_data: List[Dict[str, Any]], exchanges_data: List[Dict[str, Any]], summary: Dict[str, Any], date_range: str) -> BytesIO:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
        elements = []

        elements.append(Paragraph(settings.BUSINESS_NAME.upper(), self._get_header_style()))
        elements.append(Paragraph("Reporte de Devoluciones y Cambios", self._get_sub_header_style()))
        elements.append(Paragraph(f"Periodo: {date_range} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", self._get_sub_header_style()))
        elements.append(Spacer(1, 10))

        section_style = ParagraphStyle(
            'SectionHeaderRE',
            parent=getSampleStyleSheet()['Heading2'],
            fontSize=13,
            textColor=colors.HexColor('#1E293B'),
            spaceBefore=15,
            spaceAfter=8,
            keepWithNext=True
        )

        # 1. Financial Impact Summary
        elements.append(Paragraph("1. Resumen de Impacto Financiero", section_style))

        summary_headers = ["Devoluciones", "Total Reembolsado ($)", "Cambios", "Diferencia Neta ($)", "Pérdida Total ($)"]
        net_loss = summary.get("total_refund_usd", 0) + abs(min(0, summary.get("total_exchange_diff_usd", 0)))
        summary_row = [
            str(summary.get("total_returns", 0)),
            f"${summary.get('total_refund_usd', 0):,.2f}",
            str(summary.get("total_exchanges", 0)),
            f"${summary.get('total_exchange_diff_usd', 0):+,.2f}",
            f"${net_loss:,.2f}",
        ]

        t_summary = Table([summary_headers, summary_row], colWidths=[120, 140, 100, 140, 140])
        t_summary.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#334155')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#F8FAFC')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 11),
            ('TEXTCOLOR', (1, 1), (1, 1), colors.HexColor('#DC2626')),
            ('TEXTCOLOR', (4, 1), (4, 1), colors.HexColor('#DC2626')),
        ]))
        elements.append(t_summary)
        elements.append(Spacer(1, 20))

        # 2. Returns Table
        elements.append(Paragraph(f"2. Detalle de Devoluciones ({len(returns_data)} registros)", section_style))

        if returns_data:
            ret_headers = ["#", "Venta", "Fecha", "Productos", "Método", "Nota Crédito", "Motivo", "Total ($)"]
            ret_rows = []
            for r in returns_data:
                ret_rows.append([
                    str(r["id"]),
                    f"#{r['sale_id']}",
                    r["date"],
                    r["items"][:30],
                    r["method"],
                    r["credit_note"],
                    r["reason"][:20],
                    f"${r['total']:,.2f}",
                ])

            # Totals row
            ret_rows.append([
                "TOTAL", "", "", "", "", "", "",
                f"${summary.get('total_refund_usd', 0):,.2f}"
            ])

            col_widths = [30, 45, 85, 180, 75, 80, 100, 65]
            t_ret = Table([ret_headers] + ret_rows, colWidths=col_widths)
            t_ret.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#DC2626')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#FEF2F2')]),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                # Footer
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEE2E2')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 9),
                ('SPAN', (0, -1), (6, -1)),
            ]))
            elements.append(t_ret)
        else:
            elements.append(Paragraph("No se encontraron devoluciones en el periodo seleccionado.", ParagraphStyle('NoData', parent=getSampleStyleSheet()['Normal'], fontSize=10, textColor=colors.grey, alignment=TA_CENTER)))

        elements.append(Spacer(1, 20))

        # 3. Exchanges Table
        elements.append(Paragraph(f"3. Detalle de Cambios ({len(exchanges_data)} registros)", section_style))

        if exchanges_data:
            ex_headers = ["#", "Venta", "Fecha", "Devueltos por Cliente", "Entregados al Cliente", "Diferencia ($)", "Motivo"]
            ex_rows = []
            for ex in exchanges_data:
                diff_val = ex["difference"]
                diff_str = f"${diff_val:+,.2f}"
                ex_rows.append([
                    str(ex["id"]),
                    f"#{ex['sale_id']}",
                    ex["date"],
                    ex["items_out"][:28],
                    ex["items_in"][:28],
                    diff_str,
                    ex["reason"][:20],
                ])

            col_widths_ex = [30, 45, 85, 170, 170, 80, 80]
            t_ex = Table([ex_headers] + ex_rows, colWidths=col_widths_ex)
            t_ex.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D97706')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('TOPPADDING', (0, 0), (-1, 0), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFFBEB')]),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
            ]))
            elements.append(t_ex)
        else:
            elements.append(Paragraph("No se encontraron cambios en el periodo seleccionado.", ParagraphStyle('NoDataEx', parent=getSampleStyleSheet()['Normal'], fontSize=10, textColor=colors.grey, alignment=TA_CENTER)))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_returns_exchanges_excel(self, returns_data: List[Dict[str, Any]], exchanges_data: List[Dict[str, Any]], summary: Dict[str, Any]) -> BytesIO:
        wb = openpyxl.Workbook()
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # --- Sheet 1: Resumen ---
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        ws_summary.merge_cells('A1:D1')
        ws_summary.cell(row=1, column=1, value=settings.BUSINESS_NAME.upper())
        ws_summary.cell(row=1, column=1).font = Font(bold=True, size=16)
        ws_summary.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        summary_headers = ["Métrica", "Valor"]
        summary_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        for col, h in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = summary_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        summary_rows = [
            ("Total Devoluciones", summary.get("total_returns", 0)),
            ("Total Reembolsado ($)", summary.get("total_refund_usd", 0)),
            ("Total Cambios", summary.get("total_exchanges", 0)),
            ("Diferencia Neta Cambios ($)", summary.get("total_exchange_diff_usd", 0)),
        ]
        for i, (label, val) in enumerate(summary_rows, 4):
            ws_summary.cell(row=i, column=1, value=label).border = border
            c = ws_summary.cell(row=i, column=2, value=val)
            c.border = border
            if isinstance(val, float):
                c.number_format = '"$"#,##0.00'

        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20

        # --- Sheet 2: Devoluciones ---
        ws_ret = wb.create_sheet("Devoluciones")
        ret_headers = ["ID", "Venta", "Fecha", "Productos", "Método", "Nota Crédito", "Motivo", "Total ($)"]
        ret_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")

        for col, h in enumerate(ret_headers, 1):
            cell = ws_ret.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = ret_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row_num, r in enumerate(returns_data, 2):
            ws_ret.cell(row=row_num, column=1, value=r["id"]).border = border
            ws_ret.cell(row=row_num, column=2, value=r["sale_id"]).border = border
            ws_ret.cell(row=row_num, column=3, value=r["date"]).border = border
            ws_ret.cell(row=row_num, column=4, value=r["items"]).border = border
            ws_ret.cell(row=row_num, column=5, value=r["method"]).border = border
            ws_ret.cell(row=row_num, column=6, value=r["credit_note"]).border = border
            ws_ret.cell(row=row_num, column=7, value=r["reason"]).border = border
            c = ws_ret.cell(row=row_num, column=8, value=r["total"])
            c.number_format = '"$"#,##0.00'
            c.border = border

        for col_letter in ['A','B','C','D','E','F','G','H']:
            ws_ret.column_dimensions[col_letter].width = 18
        ws_ret.column_dimensions['D'].width = 40

        # --- Sheet 3: Cambios ---
        ws_ex = wb.create_sheet("Cambios")
        ex_headers = ["ID", "Venta", "Fecha", "Devueltos por Cliente", "Entregados al Cliente", "Diferencia ($)", "Motivo"]
        ex_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")

        for col, h in enumerate(ex_headers, 1):
            cell = ws_ex.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = ex_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row_num, ex in enumerate(exchanges_data, 2):
            ws_ex.cell(row=row_num, column=1, value=ex["id"]).border = border
            ws_ex.cell(row=row_num, column=2, value=ex["sale_id"]).border = border
            ws_ex.cell(row=row_num, column=3, value=ex["date"]).border = border
            ws_ex.cell(row=row_num, column=4, value=ex["items_out"]).border = border
            ws_ex.cell(row=row_num, column=5, value=ex["items_in"]).border = border
            c = ws_ex.cell(row=row_num, column=6, value=ex["difference"])
            c.number_format = '"$"#,##0.00'
            c.border = border
            ws_ex.cell(row=row_num, column=7, value=ex["reason"]).border = border

        for col_letter in ['A','B','C','D','E','F','G']:
            ws_ex.column_dimensions[col_letter].width = 18
        ws_ex.column_dimensions['D'].width = 35
        ws_ex.column_dimensions['E'].width = 35

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_cash_close_excel(self, cash_data: Dict[str, Any], date_str: str) -> BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cierre de Caja"
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Title
        ws.merge_cells('A1:D1')
        ws.cell(row=1, column=1, value=settings.BUSINESS_NAME.upper())
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        ws.merge_cells('A2:D2')
        ws.cell(row=2, column=1, value=f"Cierre de Caja - {date_str}")
        ws.cell(row=2, column=1).font = Font(bold=True, size=12, color="666666")
        ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

        # Section 1: Summary
        summary_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        summary_headers = ["Transacciones", "Facturación Total ($)", "Total IVA ($)", "Ganancia Neta ($)"]
        for col, h in enumerate(summary_headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = summary_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        ws.cell(row=5, column=1, value=cash_data.get("total_sales_count", 0)).border = border
        ws.cell(row=5, column=2, value=cash_data.get("total_revenue_usd", 0)).number_format = '"$"#,##0.00'
        ws.cell(row=5, column=2).border = border
        ws.cell(row=5, column=3, value=cash_data.get("total_tax_usd", 0)).number_format = '"$"#,##0.00'
        ws.cell(row=5, column=3).border = border
        c_profit = ws.cell(row=5, column=4, value=cash_data.get("total_profit_usd", 0))
        c_profit.number_format = '"$"#,##0.00'
        c_profit.font = Font(bold=True, color="10B981")
        c_profit.border = border

        # Section 2: Payment Breakdown
        pay_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        pay_headers = ["Método de Pago", "Operaciones", "Monto Recaudado", "Equiv. USD"]
        for col, h in enumerate(pay_headers, 1):
            cell = ws.cell(row=7, column=col, value=h)
            cell.font = header_font
            cell.fill = pay_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for i, pm in enumerate(cash_data.get("payment_breakdown", []), 8):
            ws.cell(row=i, column=1, value=pm.get("name", "")).border = border
            ws.cell(row=i, column=2, value=pm.get("count", 0)).border = border
            c_amt = ws.cell(row=i, column=3, value=pm.get("amount", 0))
            cur = pm.get("currency", "USD")
            c_amt.number_format = '"Bs. "#,##0.00' if cur == "VES" else '"$"#,##0.00'
            c_amt.border = border
            c_usd = ws.cell(row=i, column=4, value=pm.get("amount_usd", 0))
            c_usd.number_format = '"$"#,##0.00'
            c_usd.border = border

        for col_letter in ['A','B','C','D']:
            ws.column_dimensions[col_letter].width = 22

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_profitability_pdf(self, products_data: List[Dict[str, Any]], summary: Dict[str, Any], date_range: str) -> BytesIO:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=30)
        elements = []

        elements.append(Paragraph(settings.BUSINESS_NAME.upper(), self._get_header_style()))
        elements.append(Paragraph("Reporte de Rentabilidad y Ventas por Producto", self._get_sub_header_style()))
        elements.append(Paragraph(f"Periodo: {date_range} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", self._get_sub_header_style()))
        elements.append(Spacer(1, 10))

        section_style = ParagraphStyle(
            'SectionHeaderProfit',
            parent=getSampleStyleSheet()['Heading2'],
            fontSize=13,
            textColor=colors.HexColor('#0F766E'),
            spaceBefore=12,
            spaceAfter=6,
            keepWithNext=True
        )

        # 1. Resumen Metrics Table
        elements.append(Paragraph("1. Resumen de Desempeño Financiero", section_style))
        metrics_headers = ["Unidades Vendidas", "Ingreso Total ($)", "Costo Total ($)", "Ganancia Bruta ($)", "Margen Promedio (%)"]
        metrics_row = [
            f"{summary.get('total_qty', 0):,}",
            f"${summary.get('total_revenue', 0.0):,.2f}",
            f"${summary.get('total_cost', 0.0):,.2f}",
            f"${summary.get('total_profit', 0.0):,.2f}",
            f"{summary.get('avg_margin', 0.0):.1f}%"
        ]
        
        t_metrics = Table([metrics_headers, metrics_row], colWidths=[150, 150, 150, 150, 140])
        t_metrics.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F766E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#D1D5DB')),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#F0FDF4')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 1), (-1, 1), 11),
            ('TEXTCOLOR', (3, 1), (3, 1), colors.HexColor('#059669')),
        ]))
        elements.append(t_metrics)
        elements.append(Spacer(1, 15))

        # 2. Product Details Table
        elements.append(Paragraph("2. Rendimiento Detallado de Productos", section_style))

        # We will wrap the product names in Paragraph to avoid overflow.
        name_cell_style = ParagraphStyle(
            'ProdNameCell',
            parent=getSampleStyleSheet()['Normal'],
            fontSize=8,
            leading=9,
            wordWrap='CJK'
        )

        headers = ["Código", "Producto", "Cant. Vendida", "Costo ($)", "Precio ($)", "Ingreso Total ($)", "Ganancia ($)", "Margen (%)"]
        rows = [headers]

        for p in products_data:
            name_p = Paragraph(p.get("name", "N/A"), name_cell_style)
            margin = p.get("margin", 0.0)
            rows.append([
                p.get("barcode", "N/A"),
                name_p,
                f"{p.get('quantity', 0):,}",
                f"${p.get('cost', 0.0):,.2f}",
                f"${p.get('price', 0.0):,.2f}",
                f"${p.get('revenue', 0.0):,.2f}",
                f"${p.get('profit', 0.0):,.2f}",
                f"{margin:.1f}%"
            ])

        col_widths = [80, 230, 60, 60, 60, 90, 90, 70]
        t_prod = Table(rows, colWidths=col_widths)
        t_prod.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0369A1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
        ]))

        elements.append(t_prod)
        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_profitability_excel(self, products_data: List[Dict[str, Any]], summary: Dict[str, Any]) -> BytesIO:
        wb = openpyxl.Workbook()
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # --- Sheet 1: Resumen ---
        ws_summary = wb.active
        ws_summary.title = "Resumen de Rentabilidad"
        ws_summary.merge_cells('A1:D1')
        ws_summary.cell(row=1, column=1, value=settings.BUSINESS_NAME.upper())
        ws_summary.cell(row=1, column=1).font = Font(bold=True, size=16)
        ws_summary.cell(row=1, column=1).alignment = Alignment(horizontal="center")

        summary_headers = ["Métrica", "Valor"]
        summary_fill = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
        for col, h in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = summary_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        summary_rows = [
            ("Total Unidades Vendidas", summary.get("total_qty", 0)),
            ("Ingreso Total ($)", summary.get("total_revenue", 0.0)),
            ("Costo Total ($)", summary.get("total_cost", 0.0)),
            ("Ganancia Bruta ($)", summary.get("total_profit", 0.0)),
            ("Margen de Ganancia Promedio", f"{summary.get('avg_margin', 0.0):.2f}%"),
        ]
        for i, (label, val) in enumerate(summary_rows, 4):
            ws_summary.cell(row=i, column=1, value=label).border = border
            c = ws_summary.cell(row=i, column=2, value=val)
            c.border = border
            if isinstance(val, float):
                c.number_format = '"$"#,##0.00'
            elif isinstance(val, int):
                c.number_format = '#,##0'

        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 20

        # --- Sheet 2: Rendimiento Detallado ---
        ws_det = wb.create_sheet("Rendimiento por Producto")
        det_headers = ["Código", "Producto", "Cant. Vendida", "Costo Unit. ($)", "Precio Unit. ($)", "Ingreso Total ($)", "Ganancia ($)", "Margen (%)"]
        det_fill = PatternFill(start_color="0369A1", end_color="0369A1", fill_type="solid")

        for col, h in enumerate(det_headers, 1):
            cell = ws_det.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = det_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

        for row_num, p in enumerate(products_data, 2):
            ws_det.cell(row=row_num, column=1, value=p["barcode"]).border = border
            ws_det.cell(row=row_num, column=2, value=p["name"]).border = border
            
            c_qty = ws_det.cell(row=row_num, column=3, value=p["quantity"])
            c_qty.number_format = '#,##0'
            c_qty.border = border

            c_cost = ws_det.cell(row=row_num, column=4, value=p["cost"])
            c_cost.number_format = '"$"#,##0.00'
            c_cost.border = border

            c_price = ws_det.cell(row=row_num, column=5, value=p["price"])
            c_price.number_format = '"$"#,##0.00'
            c_price.border = border

            c_rev = ws_det.cell(row=row_num, column=6, value=p["revenue"])
            c_rev.number_format = '"$"#,##0.00'
            c_rev.border = border

            c_prof = ws_det.cell(row=row_num, column=7, value=p["profit"])
            c_prof.number_format = '"$"#,##0.00'
            c_prof.border = border

            c_marg = ws_det.cell(row=row_num, column=8, value=p["margin"]/100.0)
            c_marg.number_format = '0.0%'
            c_marg.border = border

        # Set width
        ws_det.column_dimensions['A'].width = 15
        ws_det.column_dimensions['B'].width = 35
        for col_letter in ['C','D','E','F','G','H']:
            ws_det.column_dimensions[col_letter].width = 18

        # Totals Row at the bottom of sheet 2
        last_row = len(products_data) + 1
        tot_row = last_row + 1
        ws_det.cell(row=tot_row, column=1, value="TOTALES").font = Font(bold=True)
        ws_det.cell(row=tot_row, column=1).border = border
        ws_det.cell(row=tot_row, column=2, value="").border = border

        # Sum Quantity (C)
        cell_qty = ws_det.cell(row=tot_row, column=3, value=f"=SUM(C2:C{last_row})")
        cell_qty.font = Font(bold=True)
        cell_qty.number_format = '#,##0'
        cell_qty.border = border

        ws_det.cell(row=tot_row, column=4, value="").border = border
        ws_det.cell(row=tot_row, column=5, value="").border = border

        # Sum Revenue (F)
        cell_rev = ws_det.cell(row=tot_row, column=6, value=f"=SUM(F2:F{last_row})")
        cell_rev.font = Font(bold=True)
        cell_rev.number_format = '"$"#,##0.00'
        cell_rev.border = border

        # Sum Profit (G)
        cell_prof = ws_det.cell(row=tot_row, column=7, value=f"=SUM(G2:G{last_row})")
        cell_prof.font = Font(bold=True)
        cell_prof.number_format = '"$"#,##0.00'
        cell_prof.border = border

        # Avg Margin (H)
        cell_marg = ws_det.cell(row=tot_row, column=8, value=f"=AVERAGE(H2:H{last_row})")
        cell_marg.font = Font(bold=True)
        cell_marg.number_format = '0.0%'
        cell_marg.border = border

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

report_service = ReportService()
